from openpyxl import load_workbook

SECTION_MAP = {
    '注文': 11,
    '新築': 12,
    '分譲': 21,
}

CATEGORY_MAP = {
    '完工（精算前）': 23,
    '精算完了': 24,
}

# SECTION/CATEGORY reverse maps for display
SECTION_NAMES = {v: k for k, v in SECTION_MAP.items()}
CATEGORY_NAMES = {v: k for k, v in CATEGORY_MAP.items()}


def parse_excel(file_stream):
    """Read Excel file and return (headers, records) list of dicts."""
    wb = load_workbook(file_stream, read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows())
    headers = [cell.value for cell in rows[0]]

    records = []
    for row in rows[1:]:
        values = [cell.value for cell in row]
        record = dict(zip(headers, values))
        records.append(record)

    wb.close()
    return records


def build_point_record(row):
    """Transform Excel row dict to POINT table dict.

    Returns (record_dict, skip_reason).
    If skipped, record_dict is None.
    """
    # Lat/Lon: prefer 物件, fallback to 顧客
    lat = row.get('物件緯度') or row.get('顧客緯度')
    lon = row.get('物件経度') or row.get('顧客経度')

    if not lat or not lon:
        return None, '緯度/経度なし'

    system_id = str(row.get('システムID', '') or '')
    title = str(row.get('顧客名', '') or '')[:100]
    section_id = SECTION_MAP.get(str(row.get('案件種別', '') or ''))
    category_id = CATEGORY_MAP.get(str(row.get('案件フロー', '') or ''))

    if section_id is None:
        return None, f"案件種別不明: {row.get('案件種別')}"

    postal = row.get('物件郵便番号') or row.get('顧客郵便番号')
    pref = row.get('物件都道府県') or row.get('顧客都道府県')
    addr = row.get('物件住所') or row.get('顧客現住所')

    return {
        'TITLE': title,
        'CODE': system_id,
        'POSITION_LAT': float(lat),
        'POSITION_LON': float(lon),
        'POSTALCODE': str(postal or '')[:8] or None,
        'PREFECTURE': str(pref or '')[:20] or None,
        'ADDRESS': str(addr or '')[:512] or None,
        'DETAIL_URL': f'https://andpad.jp/manager/my/orders/{system_id}' if system_id else None,
        'DESCRIPTION': str(row.get('案件備考', '') or ''),
        'CREATED_AT': row.get('案件作成日時'),
        'SECTION_ID': section_id,
        'CATEGORY_ID': category_id,
        'STAFF_RAW': str(row.get('担当者', '') or ''),
    }, None
